"""
Bu kod, belirli bir Excel dosyasındaki iki sayfa arasında veri eşleşmesi yapar ve günceller. İşte kodun işlevselliği:

1. 'openpyxl' modülü kullanarak bir Excel dosyası yükler.
2. "1" ve "3" adında iki Excel sayfası alır. Önceden "2" sayfası kullanılıyordu, ancak bu güncellenmiş versiyonda "3" sayfası kullanılıyor.
3. "1" sayfasının B sütunundaki her satırı dolaşır. Her satırdaki değeri string'e çevirir ve "None" kontrolü yapar.
4. "3" sayfasının A sütununda, "1" sayfasındaki değerle eşleşen bir değer arar.
5. Eşleşme bulunduğunda, "3" sayfasının B sütunundaki karşılık gelen değeri alır ve "1" sayfasının K sütunundaki ilgili satıra yazar.
6. Eşleşme bulunamazsa, hiçbir işlem yapılmaz.
7. Tüm değişiklikler kaydedilir.
8. Fonksiyon, verilen dosya yolundaki Excel dosyası üzerinde çalıştırılır. Bu, belirli bir dosya yoluyla örneklenmiştir.
"""

import openpyxl

def match_and_update(file_path):
    # Excel dosyasını yükle
    workbook = openpyxl.load_workbook(file_path)

    # Sayfaları al
    sheet1 = workbook['1']
    sheet3 = workbook['3']  # Güncelleme: Sayfa 3

    # Sayfa1'deki B sütunundaki her satır için
    for row1_index, row1 in enumerate(sheet1.iter_rows(min_col=2, max_col=2, min_row=2, values_only=True), start=2):
        value_to_find = str(row1[0])  # Değeri string'e çevir
        if value_to_find == "None":
            continue
        found_match = False

        # Sayfa3'deki A sütununu kontrol et (Güncelleme)
        for row3_index, row3 in enumerate(sheet3.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True), start=2):
            value_in_sheet3 = str(row3[0])  # Değeri string'e çevir
            if value_in_sheet3 == "None":
                continue

            if value_to_find == value_in_sheet3 or value_to_find.strip() == value_in_sheet3.strip():
                # Eşleşme bulundu, B sütunundaki değeri al ve Sayfa1'deki K sütununa yaz (Güncelleme)
                corresponding_value = sheet3.cell(row=row3_index, column=2).value
                sheet1.cell(row=row1_index, column=11, value=corresponding_value)
                found_match = True
                break

        if not found_match:
            # Eşleşme bulunamadı, hiçbir şey yazma
            continue

    # Değişiklikleri kaydet
    workbook.save(file_path)

# Script'i çalıştırmak için yeni dosya yolu:
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Satın Alma\\exceller\\düzgün_1.xlsx"
match_and_update(file_path)

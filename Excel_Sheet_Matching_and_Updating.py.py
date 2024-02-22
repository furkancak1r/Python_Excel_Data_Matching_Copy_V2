"""
Bu kod, belirli bir Excel dosyasındaki iki farklı sayfa arasındaki verileri eşleştirir ve günceller:

1. 'openpyxl' modülü ile belirtilen yoldaki Excel dosyası yüklenir.
2. "3" ve "4" adlı iki Excel sayfası alınır.
3. "4" sayfasının A ve B sütunlarındaki değerler, bir sözlükte anahtar-değer çifti olarak saklanır.
4. "3" sayfasının B sütunundaki her satır için:
   - Eğer B sütunundaki değer varsa ve bu değer "4" sayfasındaki sözlükte bulunuyorsa:
     - Sözlükteki karşılık gelen değer alınır.
     - Eğer "3" sayfasının K sütunu boşsa ve "4" sayfasındaki eşleşen değer boş değilse, bu değer "3" sayfasının K sütununa yazılır.
5. Tüm değişiklikler Excel dosyasına kaydedilir.

Not: Bu işlem, "3" sayfasında yer alan değerlerin "4" sayfasında eşleşen karşılıklarını bulup, bunları "3" sayfasındaki ilgili yerlere yerleştirir.
"""

import openpyxl

# Excel dosyasını yükle
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Satın Alma\\exceller\\düzgün_1.xlsx"
wb = openpyxl.load_workbook(file_path)

# Sayfaları al
sheet3 = wb["3"]
sheet4 = wb["4"]

# Sayfa 4'teki A ve B kolonundaki değerleri eşleştirerek bir sözlükte sakla
values_in_sheet4 = {sheet4.cell(row=i, column=1).value: sheet4.cell(row=i, column=2).value for i in range(1, sheet4.max_row + 1)}

# Sayfa 3'teki B kolonundaki verileri kontrol et
for row in range(1, sheet3.max_row + 1):
    value_in_b = sheet3.cell(row=row, column=2).value
    if value_in_b and value_in_b in values_in_sheet4:
        # Eğer Sayfa 4'teki eşleşen B kolonu boş değilse ve Sayfa 3'teki K kolonu boşsa, B kolonundaki değeri K'ya yaz
        corresponding_value_in_sheet4 = values_in_sheet4[value_in_b]
        if corresponding_value_in_sheet4 and not sheet3.cell(row=row, column=11).value:
            sheet3.cell(row=row, column=11).value = value_in_b

# Değişiklikleri kaydet
wb.save(file_path)

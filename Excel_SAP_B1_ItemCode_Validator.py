"""
Bu kod, bir Excel dosyasındaki verileri bir SAP Business One veritabanıyla karşılaştırır ve sonuçları günceller:

1. 'openpyxl' modülü ile belirtilen yoldaki Excel dosyası yüklenir.
2. 'pyodbc' modülü kullanılarak SAP Business One veritabanına bağlanılır. Bağlantı için gerekli parametreler (sunucu adı, veritabanı adı, kullanıcı adı ve şifre) sağlanır.
3. 'Sayfa1' isimli bir Excel sayfasındaki her satırda 'ItemCode' alınır.
4. 'ItemCode', veritabanındaki 'OITM' tablosunda sorgulanır.
5. Eğer 'ItemCode' veritabanında bulunursa, Excel dosyasındaki ilgili satırın 2. sütununa bu kod yazılır.
6. Tüm değişiklikler Excel dosyasına kaydedilir.
7. Veritabanı bağlantısı güvenli bir şekilde kapatılır.

Not: Veritabanı sorgusu için SQL sorgusu kullanılır ve 'ItemCode' değeri parametre olarak geçilir.
"""

import openpyxl
import pyodbc

# Excel dosyasını yükle
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Satın Alma\\exceller\\kontrol_edilecek.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb['Sayfa1']

# SAP Business One veritabanı bağlantısı
conn = pyodbc.connect('DRIVER={SQL Server};SERVER=server_name;DATABASE=database_name;UID=user;PWD=password')
cursor = conn.cursor()

# ItemCode'ları kontrol et ve sonuçları kaydet
for row in range(2, sheet.max_row + 1):
    item_code = sheet.cell(row=row, column=1).value
    if item_code:
        # Ensure the item_code is converted to a string
        cursor.execute("SELECT COUNT(*) FROM OITM WHERE ItemCode = ?", (str(item_code),))
        result = cursor.fetchone()
        if result[0] > 0:
            sheet.cell(row=row, column=2).value = item_code

# Değişiklikleri kaydet ve bağlantıyı kapat
wb.save(file_path)
cursor.close()
conn.close()
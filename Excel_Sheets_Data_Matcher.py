"""
openpyxl modülü kullanılarak bir Excel dosyası yüklenir.
"1" ve "2" adlı iki Excel sayfası alınır.
"1" sayfasının B sütunundaki her satır için:
Değer string'e dönüştürülür.
"None" ise döngüde bir sonraki adıma geçilir.
"2" sayfasının C sütununda eşleşme aranır:
Bu sütundaki her değer string'e dönüştürülür.
"None" ise döngüde bir sonraki adıma geçilir.
Eşleşme bulunursa:
"2" sayfasının A sütunundaki karşılık gelen değer alınır.
Bu değer "1" sayfasının K sütunundaki ilgili satıra yazılır.
Eşleşme bulundu olarak işaretlenir ve döngüden çıkılır.
Eşleşme bulunamazsa, hiçbir şey yapılmaz.
Tüm değişiklikler kaydedilir.
Fonksiyon, belirtilen dosya yolu üzerinde çalıştırılır.

"""

import openpyxl

def match_and_update(file_path):
    # Excel dosyasını yükle
    workbook = openpyxl.load_workbook(file_path)

    # Sayfaları al
    sheet1 = workbook['1']
    sheet2 = workbook['2']

    # Sayfa1'deki B sütunundaki her satır için
    for row1_index, row1 in enumerate(sheet1.iter_rows(min_col=2, max_col=2, min_row=2, values_only=True), start=2):
        value_to_find = str(row1[0])  # Değeri string'e çevir
        if value_to_find == "None":
            continue
        found_match = False

        # Sayfa2'deki C sütununu kontrol et
        for row2_index, row2 in enumerate(sheet2.iter_rows(min_col=3, max_col=3, min_row=2, values_only=True), start=2):
            value_in_sheet2 = str(row2[0])  # Değeri string'e çevir
            if value_in_sheet2 == "None":
                continue

            if value_to_find == value_in_sheet2 or value_to_find.strip() == value_in_sheet2.strip():
                # Eşleşme bulundu, A sütunundaki değeri al ve Sayfa1'deki K sütununa yaz
                corresponding_value = sheet2.cell(row=row2_index, column=1).value
                sheet1.cell(row=row1_index, column=11, value=corresponding_value)
                found_match = True
                break

        if not found_match:
            # Eşleşme bulunamadı, hiçbir şey yazma
            continue

    # Değişiklikleri kaydet
    workbook.save(file_path)

# Script'i çalıştırmak için yeni dosya yolu:
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Satın Alma\\exceller\\düzgün_1.xlsx"
match_and_update(file_path)

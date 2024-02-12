# Bu Python kodu, bir Excel dosyasındaki iki sayfa arasında veri eşleştirmesi ve güncellemesi yapar. 
# 'Sayfa1'deki C sütunundaki değerleri, 'Sayfa2'deki B sütunundaki değerlerle karşılaştırır. 
# Eşleşme bulunduğunda, 'Sayfa2'deki ilgili satırın A sütunundaki değeri 'Sayfa1'deki karşılık gelen satırın B sütununa yazar. 
# Eşleşme bulunamazsa hiçbir şey yapmaz. Ardından yapılan değişiklikleri Excel dosyasına kaydeder.
import openpyxl
def match_and_update(file_path):
    # Excel dosyasını yükle
    workbook = openpyxl.load_workbook(file_path)

    # Sayfaları al
    sheet1 = workbook['Sayfa1']
    sheet2 = workbook['Sayfa2']

    # Sayfa1'deki C sütunundaki her satır için
    for row1_index, row1 in enumerate(sheet1.iter_rows(min_col=3, max_col=3, min_row=2, values_only=True), start=2):
        value_to_find = row1[0]
        if value_to_find is None:
            continue
        found_match = False

        # Sayfa2'deki B sütununu kontrol et
        for row2_index, row2 in enumerate(sheet2.iter_rows(min_col=2, max_col=2, min_row=2, values_only=True), start=2):
            value_in_sheet2 = row2[0]
            if value_in_sheet2 is None:
                continue
            
            if value_to_find == value_in_sheet2 or value_to_find.strip() == value_in_sheet2.strip():
                # Eşleşme bulundu, A sütunundaki değeri al ve Sayfa1'deki B sütununa yaz
                corresponding_value = sheet2.cell(row=row2_index, column=1).value
                sheet1.cell(row=row1_index, column=2, value=corresponding_value)
                found_match = True
                break
        
        if not found_match:
            # Eşleşme bulunamadı, hiçbir şey yazma
            continue

    # Değişiklikleri kaydet
    workbook.save(file_path)

# Script'i çalıştırmak için:
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\SSH\\SSH Exceller\\ŞOK\\1.xlsx"
match_and_update(file_path)

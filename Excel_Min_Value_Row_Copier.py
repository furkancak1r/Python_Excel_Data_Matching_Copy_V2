"""
Bu fonksiyon, bir Excel dosyasında belirli bir kaynak sayfadan hedef sayfaya en küçük değere sahip satırları kopyalar:

1. 'openpyxl' modülü ile belirtilen yoldaki Excel dosyası yüklenir.
2. Kaynak ve hedef sayfalar belirlenir.
3. Kaynak sayfanın B sütunundaki her değer için, D sütunundaki sayısal değerle birlikte bir sözlükte saklanır.
4. Eğer bir değer tekrar ederse ve yeni sayısal değer daha küçükse, sözlükteki değer güncellenir.
5. Sözlükte saklanan en küçük değerlere sahip satırlar, hedef sayfaya kopyalanır.
6. Tüm değişiklikler Excel dosyasına kaydedilir.

Not: Bu işlem, belirli bir kriter (B sütunundaki en küçük D sütunundaki değer) temelinde verileri filtreleyip, bunları başka bir sayfaya aktarır.
"""

import openpyxl

def copy_rows_with_min_value(excel_path, source_sheet_name, target_sheet_name):
    # Excel dosyasını yükle
    workbook = openpyxl.load_workbook(excel_path)

    # Kaynak ve hedef sayfaları seç
    source_sheet = workbook[source_sheet_name]
    target_sheet = workbook[target_sheet_name]

    # B sütunundaki değerleri ve karşılık gelen D sütunundaki sayıları saklamak için bir sözlük
    values = {}

    # B sütunundaki her değer için
    for row in range(2, source_sheet.max_row + 1):  # Excel dosyaları 1'den başlar, başlık satırını atlamak için 2'den başla
        key = source_sheet[f'B{row}'].value
        value = source_sheet[f'D{row}'].value

        # Değerleri sözlüğe ekle, tekrar edenleri güncelle
        if key in values:
            if value < values[key][1]:
                values[key] = (row, value)
        else:
            values[key] = (row, value)

    # En küçük değerin bulunduğu satırları hedef sayfaya kopyala
    for row_num, _ in values.values():
        row_data = [source_sheet.cell(row=row_num, column=col).value for col in range(1, 11)]  # 1'den 10'a (A'dan J'ye)
        target_sheet.append(row_data)

    # Değişiklikleri kaydet
    workbook.save(excel_path)

# Fonksiyonu çağır
excel_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Satın Alma\\exceller\\Fiyat Listesi Çalışması\\düzgün_2.xlsx"
copy_rows_with_min_value(excel_path, '1', '2') 

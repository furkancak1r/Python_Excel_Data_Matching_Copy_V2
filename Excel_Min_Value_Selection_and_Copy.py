#openpyxl modülü kullanılarak bir Excel dosyası açılır.
#Üç Excel sayfası ("USD", "Sayfa1", "Sayfa2") işlenir. "Sayfa2" yoksa yeni bir sayfa oluşturulur.
#"USD" sayfasındaki değerler ve "Sayfa1" sayfasındaki satırlar alınır.
#"USD" sayfasından alınan her değer, "Sayfa1"deki ilgili satırlarla karşılaştırılır.
#Eşleşen satırlar arasından, 'ETKİN' veya 'PASİF' durumundaki ve en küçük "G" değerine sahip olanı seçilir.
#Seçilen satır "Sayfa2"ye eklenir.
#Değişiklikler kaydedilir.
#Fonksiyon, belirtilen dosya yolundaki Excel dosyası üzerinde çalıştırılır.

from decimal import Decimal
import openpyxl
def find_and_copy_data(excel_path):
    # Excel dosyasını aç
    workbook = openpyxl.load_workbook(excel_path)
    try_sheet = workbook["USD"]
    page1_sheet = workbook["Sayfa1"]

    # "Sayfa2" varsa kullan, yoksa oluştur
    if "Sayfa2" in workbook.sheetnames:
        page2_sheet = workbook["Sayfa2"]
    else:
        page2_sheet = workbook.create_sheet("Sayfa2")

    # TRY ve Sayfa1'den verileri al
    try_data = [cell.value for cell in try_sheet['A'] if cell.value is not None]
    page1_data = [row for row in page1_sheet.iter_rows(min_row=2, values_only=True) if row[1] is not None]

    # Eşleşmeleri kontrol et
    for try_value in try_data:
        # Eşleşen satırları bul
        matched_rows = [row for row in page1_data if try_value == row[1]]

        # En küçük G değerine sahip 'ETKİN' veya 'PASİF' satırını bul
        active_rows = [row for row in matched_rows if row[0] == "ETKİN"]
        passive_rows = [row for row in matched_rows if row[0] == "PASİF"]
        
        min_row = None
        if active_rows:
            min_row = min(active_rows, key=lambda r: Decimal(str(r[5]).replace(',', '.')) if r[5] is not None else Decimal('Infinity'))
        elif passive_rows:
            min_row = min(passive_rows, key=lambda r: Decimal(str(r[5]).replace(',', '.')) if r[5] is not None else Decimal('Infinity'))

        # Bulunan satırı Sayfa2'ye ekle
        if min_row:
            page2_sheet.append(min_row)

    # Değişiklikleri kaydet
    workbook.save(excel_path)

# Fonksiyonu çağır
find_and_copy_data("C:/Users/furkan.cakir/Desktop/FurkanPRS/Kodlar/Satın Alma/exceller/Fiyat Listesi Çalışması/min fiyat çalışma/calisma_1.xlsx")

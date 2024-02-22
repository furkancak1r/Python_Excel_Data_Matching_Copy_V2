"""
Bu kod, belirli bir Excel dosyasındaki iki sayfa arasındaki verileri karşılaştırır ve eşleşen verileri günceller:

1. 'openpyxl' modülü ile belirtilen yoldaki Excel dosyası yüklenir.
2. "4" ve "5" adında iki Excel sayfası alınır.
3. "4" sayfasının H sütunundaki her satır için:
   - Değer string'e dönüştürülür ve "None" kontrolü yapılır.
4. "5" sayfasının B sütununda, "4" sayfasındaki değerle eşleşen bir değer aranır.
5. Eşleşme bulunduğunda, "5" sayfasının A sütunundaki karşılık gelen değer alınır ve "4" sayfasının L sütunundaki ilgili satıra yazılır.
6. Tüm değişiklikler Excel dosyasına kaydedilir.

Not: Bu işlem, "4" sayfasındaki her H sütunu değerini, "5" sayfasındaki B sütunuyla karşılaştırır ve eşleşme varsa, "5" sayfasının A sütunundaki değeri "4" sayfasının L sütununa aktarır.
"""

import openpyxl

def match_and_update(file_path):
    # Excel dosyasını yükle
    workbook = openpyxl.load_workbook(file_path)

    # Sayfa 4 ve Sayfa 5'i al
    sheet4 = workbook['4']
    sheet5 = workbook['5']

    # Sayfa 4'teki H sütunundaki her satır için
    for row4_index, row4 in enumerate(sheet4.iter_rows(min_col=8, max_col=8, min_row=2, values_only=True), start=2):
        value_to_find = str(row4[0])
        if value_to_find == "None":
            continue

        # Sayfa 5'teki B sütununu kontrol et
        for row5_index, row5 in enumerate(sheet5.iter_rows(min_col=2, max_col=2, min_row=2, values_only=True), start=2):
            value_in_sheet5 = str(row5[0])
            if value_in_sheet5 == "None":
                continue

            if value_to_find == value_in_sheet5 or value_to_find.strip() == value_in_sheet5.strip():
                # Eşleşme bulundu, A sütunundaki değeri al ve Sayfa 4'teki L sütununa yaz
                corresponding_value = sheet5.cell(row=row5_index, column=1).value
                sheet4.cell(row=row4_index, column=12, value=corresponding_value)
                break

    # Değişiklikleri kaydet
    workbook.save(file_path)

# Kullanımı için dosya yolu
file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Satın Alma\\exceller\\düzgün_1.xlsx"
match_and_update(file_path)
